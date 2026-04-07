from __future__ import annotations

import json
import os
import re
import tempfile
import time
import unicodedata
from pathlib import Path
from typing import Any, Callable

import numpy as np
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
ENV_PATH = ROOT_DIR / ".env"
load_dotenv(ENV_PATH, override=True)


def _get_openai_api_key() -> str:
    load_dotenv(ENV_PATH, override=True)
    return (os.getenv("OPENAI_API_KEY") or "").strip()


OPENAI_API_KEY = _get_openai_api_key()

OPENAI_EMBEDDINGS_URL = "https://api.openai.com/v1/embeddings"
DEFAULT_EMBEDDING_MODEL = "text-embedding-3-small"
SEMANTIC_BASE_DIR = ROOT_DIR / "Build_Vector_Index" / "Semantic"
EMBEDDINGS_FILENAME = "embeddings.npy"
METADATA_FILENAME = "metadata.json"
MANIFEST_FILENAME = "manifest.json"
MAX_SEMANTIC_RESULTS = 50
DEFAULT_BATCH_SIZE = 128
DEFAULT_EMBEDDING_DTYPE = "float16"
TEXT_COLUMN_CANDIDATES = ("text")
MAX_EMBEDDING_INPUT_TOKENS = 8192
TARGET_EMBEDDING_INPUT_TOKENS = 6000
CHUNK_OVERLAP_TOKENS = 250

_SEMANTIC_INDEX_CACHE: dict[str, dict[str, Any]] = {}


def _strip_markdown_simple(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"(\*\*\*|\*\*|\*)", "", text)


def _sanitize_search_text(text: str) -> str:
    cleaned = (text or "").replace("\u00A0", " ")
    cleaned = re.sub(r"[\u200B-\u200D\u2060\uFEFF]", "", cleaned)
    cleaned = cleaned.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return re.sub(r"\s+", " ", cleaned).strip()


def _normalize_spaces(text: str) -> str:
    return _sanitize_search_text(text)


def _normalize_name(text: str) -> str:
    base = unicodedata.normalize("NFD", text or "")
    no_accents = "".join(ch for ch in base if unicodedata.category(ch) != "Mn")
    return re.sub(r"[^a-z0-9]+", "", no_accents.lower())


def _slugify(text: str) -> str:
    normalized = _normalize_name(text)
    return normalized or "semantic_index"


def to_embedding_plain_text(text: str) -> str:
    return _normalize_spaces(_strip_markdown_simple(text))


def _estimate_token_count(text: str) -> int:
    normalized = _normalize_spaces(text)
    if not normalized:
        return 0
    # Heuristica conservadora para texto em PT-BR: 1 token ~= 3 caracteres.
    return max(1, int(np.ceil(len(normalized) / 3)))


def _split_long_text(text: str, *, target_tokens: int = TARGET_EMBEDDING_INPUT_TOKENS) -> list[str]:
    normalized = _normalize_spaces(text)
    if not normalized:
        return []
    if _estimate_token_count(normalized) <= MAX_EMBEDDING_INPUT_TOKENS:
        return [normalized]

    words = normalized.split(" ")
    target_chars = max(3000, target_tokens * 3)
    overlap_words = max(8, CHUNK_OVERLAP_TOKENS // 2)
    chunks: list[str] = []
    start = 0

    while start < len(words):
        current_words: list[str] = []
        current_chars = 0
        index = start

        while index < len(words):
            word = words[index]
            added_chars = len(word) + (1 if current_words else 0)
            if current_words and current_chars + added_chars > target_chars:
                break
            current_words.append(word)
            current_chars += added_chars
            index += 1

        if not current_words:
            word = words[start]
            hard_limit = max(1000, target_chars)
            current_words = [word[:hard_limit]]
            remaining = word[hard_limit:]
            if remaining:
                words[start] = remaining
                index = start
            else:
                index = start + 1

        chunk = " ".join(current_words).strip()
        if chunk:
            chunks.append(chunk)

        if index >= len(words):
            break
        start = max(start + 1, index - overlap_words)

    return chunks or [normalized]


def _expand_rows_for_embedding(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    expanded_rows: list[dict[str, Any]] = []
    for item in rows:
        text_plain = str(item.get("text_plain") or "").strip()
        chunks = _split_long_text(text_plain)
        if len(chunks) == 1:
            expanded_rows.append(item)
            continue

        total_chunks = len(chunks)
        for chunk_index, chunk_text in enumerate(chunks, start=1):
            expanded_rows.append(
                {
                    "row": item["row"],
                    "text": chunk_text,
                    "text_plain": chunk_text,
                    "metadata": item["metadata"],
                    "source_text": item.get("text", ""),
                    "chunk_index": chunk_index,
                    "chunk_count": total_chunks,
                }
            )
    return expanded_rows


def _normalize_matrix_l2(values: np.ndarray) -> np.ndarray:
    norms = np.linalg.norm(values, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    return values / norms


def _normalize_vector_l2(values: np.ndarray) -> np.ndarray:
    norm = float(np.linalg.norm(values))
    if norm == 0:
        return values
    return values / norm


def _guess_column(headers: list[str], candidates: tuple[str, ...], fallback: str) -> str:
    normalized_map = {_normalize_name(header): header for header in headers if header}
    for candidate in candidates:
        key = _normalize_name(candidate)
        if key in normalized_map:
            return normalized_map[key]
    return fallback if fallback in headers else (headers[0] if headers else "")


def _resolve_index_paths(index_dir: Path) -> tuple[Path, Path, Path]:
    manifest_path = index_dir / MANIFEST_FILENAME
    embeddings_path = index_dir / EMBEDDINGS_FILENAME
    metadata_path = index_dir / METADATA_FILENAME

    return embeddings_path, metadata_path, manifest_path


def default_output_dir_for_workbook_name(workbook_name: str) -> Path:
    stem = Path(workbook_name or "semantic_index").stem
    return SEMANTIC_BASE_DIR / _slugify(stem)


def _index_id_from_dir(index_dir: Path) -> str:
    return index_dir.name.strip()


def _sanitize_index_id(index_id: str) -> str:
    cleaned = (index_id or "").strip()
    if not cleaned:
        raise ValueError("Parametro 'indexId' e obrigatorio.")
    if Path(cleaned).name != cleaned or cleaned in {".", ".."}:
        raise ValueError("Parametro 'indexId' invalido.")
    return cleaned


def resolve_index_dir(index_id: str) -> Path:
    safe_index_id = _sanitize_index_id(index_id)
    return SEMANTIC_BASE_DIR / safe_index_id


def _embedding_dimensions_for_model(model: str) -> int:
    normalized = (model or "").strip().lower()
    if normalized == "text-embedding-3-large":
        return 3072
    if normalized == "text-embedding-3-small":
        return 1536
    return 0


def _format_size_mb(size_bytes: float) -> str:
    size_mb = size_bytes / (1024 * 1024)
    return f"{size_mb:.1f} MB"


def _bytes_per_dtype(dtype_name: str) -> int:
    normalized = (dtype_name or "").strip().lower()
    if normalized == "float16":
        return 2
    return 4


def inspect_workbook_columns(xlsx_path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        available_sheets = list(workbook.sheetnames)
        active_sheet_name = sheet_name if sheet_name in available_sheets else available_sheets[0]
        sheet = workbook[active_sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(value).strip() if value is not None else "" for value in header_row]

        samples: list[dict[str, Any]] = []
        total_rows = 0
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not values or not any(value is not None and str(value).strip() for value in values):
                continue
            total_rows += 1
            if len(samples) < 5:
                row_map: dict[str, Any] = {"_row": row_index}
                for idx, value in enumerate(values):
                    key = headers[idx] if idx < len(headers) and headers[idx] else f"col_{idx + 1}"
                    row_map[key] = "" if value is None else str(value)
                samples.append(row_map)

        suggested_text_column = _guess_column(headers, TEXT_COLUMN_CANDIDATES, "text")
        metadata_columns = [header for header in headers if header and header != suggested_text_column]

        return {
            "sheet_name": active_sheet_name,
            "sheet_names": available_sheets,
            "headers": headers,
            "sample_rows": samples,
            "row_count": total_rows,
            "suggested_text_column": suggested_text_column,
            "metadata_columns": metadata_columns,
        }
    finally:
        workbook.close()


def read_semantic_rows_from_xlsx(
    xlsx_path: Path,
    *,
    sheet_name: str | None = None,
    text_column: str,
) -> list[dict[str, Any]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        active_sheet_name = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[active_sheet_name]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(value).strip() if value is not None else "" for value in header_row]

        rows: list[dict[str, Any]] = []
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not values:
                continue

            row_map: dict[str, Any] = {}
            for idx, value in enumerate(values):
                key = headers[idx] if idx < len(headers) and headers[idx] else f"col_{idx + 1}"
                row_map[key] = value

            text = _sanitize_search_text(str(row_map.get(text_column) or ""))
            if not text:
                continue

            rows.append(
                {
                    "row": row_index,
                    "text": text,
                    "text_plain": to_embedding_plain_text(text),
                    "metadata": {
                        key: row_map[key]
                        for key in headers
                        if key and key != text_column and key in row_map
                    },
                }
            )
        return rows
    finally:
        workbook.close()


def generate_embeddings_batch(texts: list[str], api_key: str, model: str) -> np.ndarray:
    payload = {
        "model": model or DEFAULT_EMBEDDING_MODEL,
        "input": texts,
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    response = requests.post(OPENAI_EMBEDDINGS_URL, headers=headers, json=payload, timeout=180)
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        detail = (response.text or "").strip() or str(exc)
        raise RuntimeError(f"Falha ao gerar embeddings: {detail}") from exc

    data = response.json()
    items = data.get("data") if isinstance(data, dict) else None
    if not isinstance(items, list) or len(items) != len(texts):
        raise RuntimeError("Resposta de embeddings invalida para o lote.")

    matrix = np.asarray([item.get("embedding") for item in items], dtype=np.float32)
    if matrix.ndim != 2:
        raise RuntimeError("Matriz de embeddings invalida.")
    return _normalize_matrix_l2(matrix)


def build_semantic_index(
    *,
    source_xlsx: Path,
    output_dir: Path,
    sheet_name: str | None,
    text_column: str,
    api_key: str,
    model: str = DEFAULT_EMBEDDING_MODEL,
    batch_size: int = DEFAULT_BATCH_SIZE,
    embedding_dtype: str = DEFAULT_EMBEDDING_DTYPE,
    index_label: str = "",
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> dict[str, Any]:
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY nao configurada.")
    if not source_xlsx.exists():
        raise FileNotFoundError(f"Arquivo-fonte nao encontrado: {source_xlsx}")
    if not text_column.strip():
        raise ValueError("Selecione a coluna de texto.")

    rows = read_semantic_rows_from_xlsx(
        source_xlsx,
        sheet_name=sheet_name,
        text_column=text_column,
    )
    if not rows:
        raise RuntimeError("Nenhuma linha valida encontrada para indexacao.")

    source_row_count = len(rows)
    rows = _expand_rows_for_embedding(rows)

    batch_size = max(1, int(batch_size or DEFAULT_BATCH_SIZE))
    output_dir.mkdir(parents=True, exist_ok=True)
    all_batches: list[np.ndarray] = []
    total_batches = int(np.ceil(len(rows) / batch_size))

    for batch_index, start in enumerate(range(0, len(rows), batch_size), start=1):
        batch = rows[start:start + batch_size]
        texts = [item["text_plain"] for item in batch]
        if progress_callback:
            progress_callback(batch_index, total_batches, f"Gerando embeddings do lote {batch_index}/{total_batches}...")
        all_batches.append(generate_embeddings_batch(texts, api_key=api_key, model=model))

    normalized_dtype = (embedding_dtype or DEFAULT_EMBEDDING_DTYPE).strip().lower()
    if normalized_dtype not in {"float32", "float16"}:
        raise ValueError("embedding_dtype invalido. Use float32 ou float16.")

    matrix = np.vstack(all_batches).astype(np.float16 if normalized_dtype == "float16" else np.float32, copy=False)
    embeddings_path = output_dir / EMBEDDINGS_FILENAME
    metadata_path = output_dir / METADATA_FILENAME
    manifest_path = output_dir / MANIFEST_FILENAME
    np.save(embeddings_path, matrix, allow_pickle=False)
    metadata_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    source_stat = source_xlsx.stat()
    manifest = {
        "index_label": index_label.strip() or source_xlsx.stem,
        "source_file": str(source_xlsx),
        "source_rows": source_row_count,
        "indexed_rows": len(rows),
        "sheet_name": sheet_name,
        "text_column": text_column,
        "metadata_columns": [header for header in rows[0].get("metadata", {}).keys()],
        "model": model,
        "embedding_dtype": normalized_dtype,
        "dimensions": int(matrix.shape[1]),
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S.000Z", time.gmtime()),
        "normalization": "l2",
        "source_mtime": time.strftime("%Y-%m-%dT%H:%M:%S.000Z", time.gmtime(source_stat.st_mtime)),
        "source_size": int(source_stat.st_size),
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "output_dir": str(output_dir),
        "embeddings_path": str(embeddings_path),
        "metadata_path": str(metadata_path),
        "manifest_path": str(manifest_path),
        "rows": len(rows),
        "source_rows": source_row_count,
        "dimensions": int(matrix.shape[1]),
        "embedding_dtype": normalized_dtype,
        "model": model,
    }


def list_semantic_indexes() -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    if not SEMANTIC_BASE_DIR.exists():
        return results

    for child in sorted((item for item in SEMANTIC_BASE_DIR.iterdir() if item.is_dir()), key=lambda item: item.name.lower()):
        embeddings_path, metadata_path, manifest_path = _resolve_index_paths(child)
        if not embeddings_path.exists() or not metadata_path.exists() or not manifest_path.exists():
            continue
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            continue

        results.append(
            {
                "id": _index_id_from_dir(child),
                "label": str(manifest.get("index_label") or child.name).strip() or child.name,
                "sourceFile": str(manifest.get("source_file") or ""),
                "sourceRows": int(manifest.get("source_rows") or 0),
                "model": str(manifest.get("model") or DEFAULT_EMBEDDING_MODEL),
                "dimensions": int(manifest.get("dimensions") or 0),
                "embeddingDtype": str(manifest.get("embedding_dtype") or DEFAULT_EMBEDDING_DTYPE),
            }
        )

    return results


def _load_semantic_index(index_id: str) -> dict[str, Any]:
    safe_index_id = _sanitize_index_id(index_id)
    cached = _SEMANTIC_INDEX_CACHE.get(safe_index_id)
    if cached is not None:
        return cached

    index_dir = resolve_index_dir(safe_index_id)
    embeddings_path, metadata_path, manifest_path = _resolve_index_paths(index_dir)
    if not embeddings_path.exists() or not metadata_path.exists() or not manifest_path.exists():
        raise FileNotFoundError(
            f"Indice semantico '{safe_index_id}' nao encontrado em {index_dir}."
        )

    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    embeddings = np.load(embeddings_path, allow_pickle=False)

    if embeddings.ndim != 2:
        raise ValueError("Arquivo de embeddings invalido: matriz esperada.")
    if not isinstance(metadata, list):
        raise ValueError("Arquivo de metadata invalido: lista esperada.")
    if len(metadata) != embeddings.shape[0]:
        raise ValueError("Indice semantico inconsistente: quantidade de embeddings difere da metadata.")

    dimensions = int(manifest.get("dimensions") or 0)
    if dimensions and dimensions != int(embeddings.shape[1]):
        raise ValueError("Indice semantico inconsistente: dimensoes do manifest nao conferem.")

    indexed_rows = int(manifest.get("indexed_rows") or manifest.get("source_rows") or 0)
    if indexed_rows and indexed_rows != len(metadata):
        raise ValueError("Indice semantico inconsistente: indexed_rows nao confere com a metadata.")

    normalization = str(manifest.get("normalization") or "").strip().lower()
    if normalization != "l2":
        embeddings = _normalize_matrix_l2(embeddings)
    embeddings = embeddings.astype(np.float32, copy=False)

    loaded_index = {
        "metadata": metadata,
        "manifest": manifest,
        "embeddings": embeddings,
    }
    _SEMANTIC_INDEX_CACHE[safe_index_id] = loaded_index
    return loaded_index


def _embed_query(text: str, api_key: str, model: str) -> np.ndarray:
    matrix = generate_embeddings_batch([text], api_key=api_key, model=model)
    return _normalize_vector_l2(matrix[0])


def search_semantic_index(
    index_id: str,
    query: str,
    limit: int = 10,
    api_key: str | None = None,
) -> tuple[int, list[dict[str, Any]]]:
    raw_query = _sanitize_search_text(query)
    if not raw_query:
        raise ValueError("Parametro 'query' e obrigatorio.")

    api_key = (api_key or _get_openai_api_key()).strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY nao configurada para Semantic Search.")

    max_rows = max(1, min(int(limit or 10), MAX_SEMANTIC_RESULTS))
    index = _load_semantic_index(index_id)
    metadata: list[dict[str, Any]] = index["metadata"]
    manifest: dict[str, Any] = index["manifest"]
    embeddings: np.ndarray = index["embeddings"]
    index_label = str(manifest.get("index_label") or _sanitize_index_id(index_id)).strip() or _sanitize_index_id(index_id)

    model = str(manifest.get("model") or DEFAULT_EMBEDDING_MODEL).strip() or DEFAULT_EMBEDDING_MODEL
    query_vector = _embed_query(to_embedding_plain_text(raw_query), api_key=api_key, model=model)
    scores = embeddings @ query_vector

    if scores.size == 0:
        return 0, []

    top_indices = np.argsort(scores)[::-1][:max_rows]
    matches: list[dict[str, Any]] = []
    for idx in top_indices.tolist():
        item = metadata[idx] if idx < len(metadata) else {}
        item_metadata = item.get("metadata")
        if not isinstance(item_metadata, dict):
            raise ValueError(
                "Indice semantico em formato antigo. Reindexe a base para usar o modelo atual de metadata."
            )
        matches.append(
            {
                "book": index_label,
                "index_id": _sanitize_index_id(index_id),
                "index_label": index_label,
                "row": int(item.get("row") or 0),
                "text": str(item.get("text") or "").strip(),
                "metadata": item_metadata,
                "score": float(scores[idx]),
            }
        )

    return len(matches), matches
def run_streamlit_app() -> None:
    import streamlit as st

    st.set_page_config(
        page_title="Semantic Index Builder",
        page_icon=":sparkles:",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown(
        """
        <style>
          .stApp {
            background:
              radial-gradient(circle at top left, rgba(16,185,129,.10), transparent 32%),
              radial-gradient(circle at top right, rgba(59,130,246,.10), transparent 28%),
              linear-gradient(180deg, #f6fbff 0%, #f8fafc 100%);
          }
          .block-container { padding-top: 1.5rem; padding-bottom: 2rem; max-width: 1220px; }
          .hero {
            padding: 1.4rem 1.6rem;
            border-radius: 24px;
            background: linear-gradient(135deg, rgba(255,255,255,.96), rgba(239,246,255,.92));
            border: 1px solid rgba(148,163,184,.22);
            box-shadow: 0 24px 64px -36px rgba(15,23,42,.35);
            margin-bottom: 1rem;
          }
          .hero h1 { margin: 0; font-size: 2rem; color: #0f172a; }
          .hero p { margin: .4rem 0 0; color: #475569; font-size: .98rem; }
          .card {
            background: rgba(255,255,255,.92);
            border: 1px solid rgba(148,163,184,.22);
            border-radius: 22px;
            padding: 1rem 1.1rem;
            box-shadow: 0 20px 52px -38px rgba(15,23,42,.32);
          }
          .metric {
            padding: .9rem 1rem;
            border-radius: 18px;
            background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
            border: 1px solid rgba(148,163,184,.18);
          }
          .metric-label { font-size: .76rem; text-transform: uppercase; letter-spacing: .08em; color: #64748b; }
          .metric-value { font-size: 1.22rem; font-weight: 700; color: #0f172a; }
          .small-note { color: #64748b; font-size: .84rem; }
        </style>
        <div class="hero">
          <h1>Semantic Index Builder</h1>
          <p>Upload de planilhas, mapeamento flexivel de colunas e geracao de artefatos vetoriais prontos para uso.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if "workbook_bytes" not in st.session_state:
        st.session_state.workbook_bytes = None
    if "workbook_name" not in st.session_state:
        st.session_state.workbook_name = ""
    if "inspection" not in st.session_state:
        st.session_state.inspection = None
    if "selected_text_column" not in st.session_state:
        st.session_state.selected_text_column = ""

    with st.sidebar:
        st.markdown("### Configuracao")
        api_key = st.text_input(
            "OPENAI API Key",
            value=_get_openai_api_key(),
            type="password",
            help="Usada para gerar os embeddings.",
        ).strip()
        model = st.selectbox(
            "Modelo",
            options=["text-embedding-3-large", "text-embedding-3-small"],
            index=["text-embedding-3-large", "text-embedding-3-small"].index(DEFAULT_EMBEDDING_MODEL),
        )
        embedding_dtype = st.selectbox(
            "Tipo do arquivo de embeddings",
            options=["float32", "float16"],
            index=["float32", "float16"].index(DEFAULT_EMBEDDING_DTYPE),
            help="float16 reduz o tamanho do arquivo aproximadamente pela metade, com pequena perda de precisao.",
        )
        batch_size = st.slider("Batch size", min_value=16, max_value=256, value=DEFAULT_BATCH_SIZE, step=16)
        st.caption("A pasta de saida sugerida segue o nome do arquivo Excel sem a extensao.")

    upload_col, info_col = st.columns([1.35, 0.65], gap="large")

    with upload_col:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        uploaded = st.file_uploader("Planilha Excel (.xlsx)", type=["xlsx"])
        if uploaded is not None:
            workbook_bytes = uploaded.getvalue()
            if workbook_bytes != st.session_state.workbook_bytes:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                    temp_file.write(workbook_bytes)
                    temp_path = Path(temp_file.name)
                try:
                    inspection = inspect_workbook_columns(temp_path)
                finally:
                    temp_path.unlink(missing_ok=True)
                st.session_state.workbook_bytes = workbook_bytes
                st.session_state.workbook_name = uploaded.name
                st.session_state.inspection = inspection

        inspection = st.session_state.inspection
        if inspection is None:
            st.info("Envie uma planilha para inspecionar as colunas e gerar o indice.")
        else:
            sheet_names = inspection["sheet_names"]
            headers = inspection["headers"]
            st.markdown(f"**Arquivo:** `{st.session_state.workbook_name}`")
            selected_sheet = st.selectbox("Sheet", options=sheet_names, index=sheet_names.index(inspection["sheet_name"]))
            if selected_sheet != inspection["sheet_name"] and st.session_state.workbook_bytes:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                    temp_file.write(st.session_state.workbook_bytes)
                    temp_path = Path(temp_file.name)
                try:
                    inspection = inspect_workbook_columns(temp_path, sheet_name=selected_sheet)
                finally:
                    temp_path.unlink(missing_ok=True)
                st.session_state.inspection = inspection
                headers = inspection["headers"]

            if st.session_state.selected_text_column not in headers:
                st.session_state.selected_text_column = (
                    inspection["suggested_text_column"] if inspection["suggested_text_column"] in headers else (headers[0] if headers else "")
                )

            text_column = st.selectbox(
                "Coluna de texto",
                options=headers,
                index=headers.index(st.session_state.selected_text_column) if st.session_state.selected_text_column in headers else 0,
            )
            st.session_state.selected_text_column = text_column
            metadata_columns = [header for header in headers if header and header != text_column]

            st.markdown("#### Metadados")
            if metadata_columns:
                st.caption("As colunas abaixo serao gravadas como metadados no `metadata.json`.")
                st.markdown(", ".join(f"`{column}`" for column in metadata_columns))
            else:
                st.caption("Nenhuma outra coluna disponivel para metadados.")

            default_output_dir = default_output_dir_for_workbook_name(st.session_state.workbook_name)
            output_dir_text = st.text_input("Diretorio de saida", value=str(default_output_dir)).strip() or str(default_output_dir)
            index_label = st.text_input("Rotulo do indice", value=Path(st.session_state.workbook_name).stem)

            st.markdown("#### Preview")
            preview_rows = inspection["sample_rows"]
            if preview_rows:
                st.dataframe(preview_rows, use_container_width=True, hide_index=True)
            else:
                st.warning("Nenhuma linha valida encontrada na previa.")

            progress = st.progress(0, text="Aguardando geracao...")
            status_box = st.empty()

            def _progress_callback(batch_index: int, total_batches: int, message: str) -> None:
                percent = int((batch_index / max(total_batches, 1)) * 100)
                progress.progress(percent, text=message)
                status_box.info(message)

            if st.button("Gerar embeddings", type="primary", use_container_width=True):
                if not api_key:
                    st.error("Informe a OPENAI API Key para gerar os embeddings.")
                elif not st.session_state.workbook_bytes:
                    st.error("Envie uma planilha antes de gerar o indice.")
                else:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                        temp_file.write(st.session_state.workbook_bytes)
                        temp_path = Path(temp_file.name)
                    try:
                        result = build_semantic_index(
                            source_xlsx=temp_path,
                            output_dir=Path(output_dir_text),
                            sheet_name=inspection["sheet_name"],
                            text_column=text_column,
                            api_key=api_key,
                            model=model,
                            batch_size=batch_size,
                            embedding_dtype=embedding_dtype,
                            index_label=index_label,
                            progress_callback=_progress_callback,
                        )
                    except Exception as exc:
                        progress.progress(0, text="Falha na geracao.")
                        status_box.error(str(exc))
                        st.exception(exc)
                    else:
                        progress.progress(100, text="Indice gerado com sucesso.")
                        status_box.success("Artefatos gerados com sucesso.")
                        st.success("Embeddings gerados.")
                        st.json(result)
                    finally:
                        temp_path.unlink(missing_ok=True)
        st.markdown("</div>", unsafe_allow_html=True)

    with info_col:
        inspection = st.session_state.inspection
        dimensions = _embedding_dimensions_for_model(model)
        dtype_bytes = _bytes_per_dtype(embedding_dtype)
        estimated_embeddings_bytes = (
            float(inspection["row_count"] * dimensions * dtype_bytes) if inspection is not None and dimensions > 0 else 0.0
        )
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### Resumo")
        if inspection is None:
            st.markdown('<div class="small-note">A previa da planilha aparecera aqui.</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                f"""
                <div class="metric">
                  <div class="metric-label">Linhas validas</div>
                  <div class="metric-value">{inspection['row_count']}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
                <div class="metric" style="margin-top:.8rem;">
                  <div class="metric-label">Sheets</div>
                  <div class="metric-value">{len(inspection['sheet_names'])}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
                <div class="metric" style="margin-top:.8rem;">
                  <div class="metric-label">Colunas detectadas</div>
                  <div class="metric-value">{len(inspection['headers'])}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
                <div class="metric" style="margin-top:.8rem;">
                  <div class="metric-label">Dimensao do vetor</div>
                  <div class="metric-value">{dimensions}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
                <div class="metric" style="margin-top:.8rem;">
                  <div class="metric-label">Tipo numerico</div>
                  <div class="metric-value">{embedding_dtype}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown(
                f"""
                <div class="metric" style="margin-top:.8rem;">
                  <div class="metric-label">Estimativa do embeddings.npy</div>
                  <div class="metric-value">{_format_size_mb(estimated_embeddings_bytes)}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            st.markdown("### Colunas")
            for header in inspection["headers"]:
                st.markdown(f"- `{header or '(vazia)'}`")
            st.markdown("### Metadados")
            selected_text_column = st.session_state.selected_text_column or inspection["suggested_text_column"]
            metadata_columns = [header for header in inspection["headers"] if header and header != selected_text_column]
            if metadata_columns:
                for header in metadata_columns:
                    st.markdown(f"- `{header}`")
            else:
                st.markdown("- Nenhuma")

        st.markdown("### Saida")
        st.markdown(
            """
            - `embeddings.npy`
            - `metadata.json`
            - `manifest.json`
            """,
        )
        if inspection is not None:
            st.caption(
                "A estimativa considera apenas o arquivo embeddings.npy "
                f"({inspection['row_count']} linhas x {dimensions} dimensoes x {dtype_bytes} bytes)."
            )
        st.caption("O gerador aceita pequenas variacoes nos nomes das colunas e permite selecao manual.")
        st.markdown("</div>", unsafe_allow_html=True)


if __name__ == "__main__":
    run_streamlit_app()
