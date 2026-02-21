from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any, Callable, Optional

try:
    import openpyxl  # type: ignore
except Exception as exc:  # pragma: no cover - import guard
    raise RuntimeError("Dependency 'openpyxl' is required for lexical search.") from exc


LEXICAL_DIR = Path(__file__).resolve().parents[1] / "Files" / "Lexical"
_BOOL_OPS: dict[str, int] = {"!": 3, "&": 2, "|": 1}
BOOK_CODE_TO_FILE: dict[str, str] = {
    "PC": "PC",  # nao ha arquivo na base atual
    "PROJ": "PROJ",
    "EXP": "700EXP",
    "CCG": "CCG",
    "TNP": "TNP",
    "MP": "PROEXIS",
    "MDE": "DUPLA",
    "MRC": "MRC",  # nao ha arquivo na base atual
    "MMT": "MMT",  # nao ha arquivo na base atual
    "TC": "TEMAS",
    "TEAT": "200TEAT",
    "NE": "NE",  # nao ha arquivo na base atual
    "HSR": "HSR",
    "HSP": "HSP",
    "DNC": "DNC",  # nao ha arquivo na base atual
    "DAC": "DAC",
    "LO": "LO",
    "EC": "EC",
}
FILE_TO_BOOK_CODE: dict[str, str] = {filename: code for code, filename in BOOK_CODE_TO_FILE.items()}

# Limite proprio do backend para buscas lexicais em livro.
# A varredura e interrompida quando atingir esse teto para evitar
# processamento excessivo em arquivos muito grandes.
MAX_BOOK_SEARCH = 200


def _normalize(text: str) -> str:
    base = unicodedata.normalize("NFD", text or "")
    return "".join(ch for ch in base if unicodedata.category(ch) != "Mn").lower().strip()


def _normalize_for_match(text: str) -> str:
    base = _normalize(text or "")
    return re.sub(r"[^\w\s\*]", "", base)


def _strip_markdown_simple(text: str) -> str:
    """
    Remove marcacoes markdown simples (* e **) para evitar ruido de match.
    """
    if not text:
        return ""
    return re.sub(r"(\*\*|\*)", "", text)


def _balanced_parentheses(query: str) -> bool:
    depth = 0
    for ch in query:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth < 0:
                return False
    return depth == 0


def _tokenize_query(query: str) -> list[str]:
    tokens: list[str] = []
    i, n = 0, len(query)
    while i < n:
        c = query[i]
        if c.isspace():
            i += 1
            continue
        if c in "()&|!":
            tokens.append(c)
            i += 1
            continue
        if c == '"':
            j = i + 1
            buf: list[str] = []
            while j < n and query[j] != '"':
                buf.append(query[j])
                j += 1
            tokens.append('"' + "".join(buf) + '"')
            i = j + 1 if j < n and query[j] == '"' else j
            continue
        j = i
        while j < n and (not query[j].isspace()) and query[j] not in "()&|!":
            j += 1
        tokens.append(query[i:j])
        i = j
    return tokens


def _shunting_yard(tokens: list[str]) -> list[str]:
    output: list[str] = []
    st: list[str] = []
    for t in tokens:
        if t in _BOOL_OPS:
            while st and st[-1] in _BOOL_OPS and _BOOL_OPS[st[-1]] >= _BOOL_OPS[t]:
                output.append(st.pop())
            st.append(t)
        elif t == "(":
            st.append(t)
        elif t == ")":
            while st and st[-1] != "(":
                output.append(st.pop())
            if st and st[-1] == "(":
                st.pop()
        else:
            output.append(t)
    while st:
        output.append(st.pop())
    return output


def _wildcard_pattern(term_raw: str) -> re.Pattern[str]:
    term = _normalize_for_match(term_raw)
    body = "".join(".*" if ch == "*" else re.escape(ch) for ch in term)
    return re.compile(rf"\b{body}\b", flags=re.IGNORECASE)


def _phrase_pattern(quoted_raw: str) -> re.Pattern[str]:
    core = quoted_raw[1:-1] if len(quoted_raw) >= 2 and quoted_raw[0] == '"' and quoted_raw[-1] == '"' else quoted_raw
    core_norm = _normalize_for_match(core)
    parts = [re.escape(p) for p in core_norm.split() if p]
    if not parts:
        return re.compile(r"(?!x)x")
    return re.compile(r"\b" + r"\s+".join(parts) + r"\b", flags=re.IGNORECASE)


def _compile_boolean_predicate(query: str) -> Callable[[str], bool]:
    q = (query or "").strip()
    if not q:
        return lambda _pnorm: True

    tokens = _tokenize_query(q)
    if not tokens:
        return lambda _pnorm: True
    if not _balanced_parentheses(q):
        return lambda pnorm: _normalize_for_match(q) in pnorm

    rpn = _shunting_yard(tokens)
    pat_cache: dict[str, re.Pattern[str] | str] = {}

    for token in tokens:
        if token in _BOOL_OPS or token in {"(", ")"}:
            continue
        if token.startswith('"') and token.endswith('"') and len(token) >= 2:
            pat_cache[token] = _phrase_pattern(token)
        elif "*" in token:
            pat_cache[token] = _wildcard_pattern(token)
        else:
            pat_cache[token] = _normalize_for_match(token)

    def _eval_token(token: str, pnorm: str) -> bool:
        compiled = pat_cache.get(token)
        if compiled is None:
            return False
        if isinstance(compiled, str):
            return bool(compiled) and compiled in pnorm
        return compiled.search(pnorm) is not None

    def _predicate(pnorm: str) -> bool:
        st: list[bool] = []
        for t in rpn:
            if t in _BOOL_OPS:
                if t == "!":
                    if not st:
                        return False
                    st.append(not st.pop())
                else:
                    if len(st) < 2:
                        return False
                    b = st.pop()
                    a = st.pop()
                    st.append(a and b if t == "&" else a or b)
            else:
                st.append(_eval_token(t, pnorm))
        return len(st) == 1 and st[0]

    return _predicate


def _compile_prefilter(query: str) -> Optional[Callable[[str], bool]]:
    q = (query or "").strip()
    if not q:
        return None

    tokens = _tokenize_query(q)
    if not tokens or "|" in tokens:
        return None

    must_have: list[str] = []
    must_not: list[str] = []
    negate_next = False

    for token in tokens:
        if token in {"(", ")", "&"}:
            continue
        if token == "!":
            negate_next = True
            continue
        if token == "|":
            return None
        if "*" in token:
            negate_next = False
            continue
        core = token[1:-1] if token.startswith('"') and token.endswith('"') and len(token) >= 2 else token
        lit = _normalize_for_match(core)
        if not lit:
            negate_next = False
            continue
        if negate_next:
            must_not.append(lit)
            negate_next = False
        else:
            must_have.append(lit)

    if not must_have and not must_not:
        return None

    def _prefilter(pnorm: str) -> bool:
        return all(lit in pnorm for lit in must_have) and all(lit not in pnorm for lit in must_not)

    return _prefilter


def _process_found_paragraph(paragraph: str, search_term: str) -> str:
    """
    Reestrutura paragrafos que usam '|' como agregador:
      - Se houver 2+ ocorrencias de '|': mantem o primeiro "cabecalho" e
        adiciona apenas subtrechos que contenham o termo de busca (normalizado).
      - Caso contrario, retorna o paragrafo original.
    """
    if not search_term:
        return paragraph

    needle = _normalize_for_match(search_term)
    if not needle:
        return paragraph

    if paragraph.count("|") >= 2:
        parts = paragraph.split("|")
        if not parts:
            return paragraph

        rebuilt: list[str] = [parts[0].strip()]
        for sub in parts[1:]:
            segment = sub.strip()
            if needle in _normalize_for_match(segment):
                rebuilt.append(segment)

        if len(rebuilt) == 1:
            return ""

        result = " ".join(rebuilt)
        return result.replace("|", "").replace("\\", "").replace("\n", "").strip()

    return paragraph


def list_lexical_books() -> list[str]:
    if not LEXICAL_DIR.exists():
        return []
    codes: list[str] = []
    for code, filename in BOOK_CODE_TO_FILE.items():
        if not (LEXICAL_DIR / f"{filename}.xlsx").exists():
            continue
        codes.append(code)
    return sorted(codes)


def _search_lexical_book_internal(book: str, term: str, limit: int = 50) -> tuple[int, list[dict[str, Any]]]:
    book_code = (book or "").strip().upper()
    raw_term = (term or "").strip()
    if not book_code:
        raise ValueError("Parametro 'book' e obrigatorio.")
    if not raw_term:
        raise ValueError("Parametro 'term' e obrigatorio.")

    filename = BOOK_CODE_TO_FILE.get(book_code)
    if not filename:
        # Compatibilidade legada: aceita nome de arquivo/stem diretamente.
        filename = (book or "").strip()
    source_path = LEXICAL_DIR / f"{filename}.xlsx"
    if not source_path.exists():
        raise FileNotFoundError(f"Livro lexical nao encontrado: {book_code}.")

    max_rows = max(1, min(int(limit or 50), MAX_BOOK_SEARCH))
    term_norm = _normalize_for_match(raw_term)
    if not term_norm:
        raise ValueError("Parametro 'term' invalido.")
    predicate = _compile_boolean_predicate(raw_term)
    prefilter = _compile_prefilter(raw_term)

    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [str(col).strip().lower() if col is not None else "" for col in header_row]

        rows: list[dict[str, Any]] = []
        total_matches = 0
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not values:
                continue
            row_map: dict[str, Any] = {}
            for idx, value in enumerate(values):
                key = headers[idx] if idx < len(headers) and headers[idx] else f"col_{idx + 1}"
                row_map[key] = "" if value is None else str(value).strip()

            normalized_values: list[str] = []
            for value in row_map.values():
                if not value:
                    continue
                cleaned = _strip_markdown_simple(str(value))
                normalized_values.append(cleaned)
            haystack = _normalize_for_match(" ".join(normalized_values))
            if prefilter is not None and not prefilter(haystack):
                continue
            if not predicate(haystack):
                continue

            row_number_raw = row_map.get("number") or row_map.get("paragraph_number") or ""
            try:
                row_number = int(str(row_number_raw))
            except Exception:
                row_number = None

            raw_text = _strip_markdown_simple(str(row_map.get("text") or "").strip())
            processed_text = _process_found_paragraph(raw_text, raw_term) if raw_text else raw_text
            if raw_text and not processed_text:
                continue

            total_matches += 1
            if len(rows) < max_rows:
                rows.append(
                    {
                        "book": book_code,
                        "row": row_index,
                        "number": row_number,
                        "title": str(row_map.get("title") or "").strip(),
                        "text": processed_text,
                        "data": row_map,
                    }
                )
            if total_matches >= MAX_BOOK_SEARCH:
                break

        return total_matches, rows
    finally:
        workbook.close()


def search_lexical_book(book: str, term: str, limit: int = 50) -> list[dict[str, Any]]:
    _, rows = _search_lexical_book_internal(book=book, term=term, limit=limit)
    return rows


def search_lexical_book_with_total(book: str, term: str, limit: int = 50) -> tuple[int, list[dict[str, Any]]]:
    return _search_lexical_book_internal(book=book, term=term, limit=limit)


def search_lexical_verbetes_with_total(
    author: str = "",
    title: str = "",
    area: str = "",
    text: str = "",
    limit: int = 50,
) -> tuple[int, list[dict[str, Any]]]:
    source_path = LEXICAL_DIR / "EC.xlsx"
    if not source_path.exists():
        raise FileNotFoundError("Base de verbetes nao encontrada: EC.xlsx")

    filters = {
        "author": (author or "").strip(),
        "title": (title or "").strip(),
        "area": (area or "").strip(),
        "text": (text or "").strip(),
    }
    active_filters = {key: value for key, value in filters.items() if value}
    if not active_filters:
        raise ValueError("Informe ao menos um campo de busca.")

    normalized_filters = {key: _normalize_for_match(value) for key, value in active_filters.items()}
    max_rows = max(1, min(int(limit or 50), MAX_BOOK_SEARCH))

    # Para verbetes, precisamos acessar hyperlinks das celulas (coluna "link"),
    # por isso usamos modo normal (nao read_only).
    workbook = openpyxl.load_workbook(source_path, read_only=False, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=False), ())
        headers = [str(cell.value).strip().lower() if getattr(cell, "value", None) is not None else "" for cell in header_cells]

        rows: list[dict[str, Any]] = []
        total_matches = 0
        for row_index, cells in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
            if not cells:
                continue

            row_map: dict[str, Any] = {}
            for idx, cell in enumerate(cells):
                value = getattr(cell, "value", None)
                key = headers[idx] if idx < len(headers) and headers[idx] else f"col_{idx + 1}"
                cell_text = "" if value is None else str(value).strip()
                if key == "link":
                    hyperlink = getattr(cell, "hyperlink", None)
                    target = ""
                    if hyperlink is not None:
                        target = str(getattr(hyperlink, "target", "") or getattr(hyperlink, "location", "") or "").strip()
                    row_map[key] = target or cell_text
                else:
                    row_map[key] = cell_text

            matched = True
            for field, needle in normalized_filters.items():
                field_value = _normalize_for_match(str(row_map.get(field) or ""))
                if not needle or needle not in field_value:
                    matched = False
                    break
            if not matched:
                continue

            row_number_raw = row_map.get("number") or ""
            try:
                row_number = int(str(row_number_raw))
            except Exception:
                row_number = None

            total_matches += 1
            if len(rows) < max_rows:
                rows.append(
                    {
                        "row": row_index,
                        "number": row_number,
                        "title": str(row_map.get("title") or "").strip(),
                        "text": str(row_map.get("text") or "").strip(),
                        "link": str(row_map.get("link") or "").strip(),
                        "data": row_map,
                    }
                )
            if total_matches >= MAX_BOOK_SEARCH:
                break

        return total_matches, rows
    finally:
        workbook.close()
