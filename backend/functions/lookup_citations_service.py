from __future__ import annotations

import pickle
import re
import unicodedata
from bisect import bisect_left, bisect_right
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from threading import Lock
from typing import Any, Iterable

try:
    import openpyxl  # type: ignore
except Exception as exc:  # pragma: no cover - import guard
    raise RuntimeError("Dependency 'openpyxl' is required for lexical citation lookup.") from exc

from rapidfuzz import fuzz


BASE_DIR = Path(__file__).resolve().parents[1]
LEXICAL_DIR = BASE_DIR / "Files" / "Lexical"
INDEX_CACHE_PATH = Path(__file__).resolve().parent / ".lexical_index.pkl"
INDEX_CACHE_VERSION = 2
STOPWORDS = {
    "a",
    "as",
    "ao",
    "aos",
    "com",
    "da",
    "das",
    "de",
    "do",
    "dos",
    "e",
    "em",
    "na",
    "nas",
    "no",
    "nos",
    "o",
    "os",
    "ou",
    "para",
    "por",
    "que",
    "se",
    "sem",
    "um",
    "uma",
}
MAX_TOKENS_DOCUMENTO = 60
MAX_TOKENS_CONSULTA = 12
MAX_CANDIDATOS_GLOBAIS = 1200
SCORE_MINIMO_FALLBACK = 70

PROCESS_INDEX_CACHE: dict[str, Any] = {"manifesto": None, "indice": None}
PROCESS_INDEX_CACHE_LOCK = Lock()


@dataclass(slots=True)
class LexicalEntry:
    arquivo: str
    pagina: int | None
    texto: str
    texto_norm: str
    alnum_len: int
    ordem: int
    referencia_contexto: int


def normalizar(texto: Any) -> str:
    texto = "" if texto is None else str(texto)
    texto = texto.lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def tokenizar(texto: str, limite: int) -> list[str]:
    tokens: list[str] = []
    vistos: set[str] = set()

    for token in re.findall(r"[a-z0-9]+", texto):
        if len(token) < 4 or token in STOPWORDS or token in vistos:
            continue
        vistos.add(token)
        tokens.append(token)
        if len(tokens) >= limite:
            break

    return tokens


def contar_alnum(texto: str) -> int:
    return sum(1 for char in texto if char.isalnum())


def valor_ausente(valor: Any) -> bool:
    if valor is None:
        return True
    return isinstance(valor, float) and valor != valor


def pagina_valida(valor: Any) -> int | None:
    if valor_ausente(valor):
        return None
    try:
        return int(valor)
    except (TypeError, ValueError):
        try:
            return int(float(valor))
        except (TypeError, ValueError):
            return None


def criar_resultado_vazio() -> dict[str, Any]:
    return {
        "pagina": None,
        "score": 0,
        "metodo": None,
        "paragrafo_lexical": "",
        "_arquivo": None,
        "_referencia_contexto": None,
        "_ordem": None,
    }


def criar_resultado(entrada: LexicalEntry, score: float, metodo: str) -> dict[str, Any]:
    return {
        "pagina": entrada.pagina,
        "score": score,
        "metodo": metodo,
        "paragrafo_lexical": entrada.texto,
        "_arquivo": entrada.arquivo,
        "_referencia_contexto": entrada.referencia_contexto,
        "_ordem": entrada.ordem,
    }


def resultado_para_saida(paragrafo_entrada: str, resultado: dict[str, Any]) -> dict[str, Any]:
    return {
        "inputParagraph": paragrafo_entrada,
        "matchedParagraph": resultado["paragrafo_lexical"],
        "book": resultado["_arquivo"] or "N/D",
        "page": resultado["pagina"] if resultado["pagina"] is not None else "N/D",
        "similarity": round(float(resultado["score"]), 2),
        "method": resultado["metodo"] or "sem_match",
        "matchedRow": resultado["_ordem"] if resultado["_ordem"] is not None else "N/D",
        "matchedReference": resultado["_referencia_contexto"] if resultado["_referencia_contexto"] is not None else "N/D",
    }


def coletar_manifesto_lexical() -> list[dict[str, Any]]:
    manifesto: list[dict[str, Any]] = []

    for caminho in sorted(LEXICAL_DIR.glob("*.xlsx")):
        stat = caminho.stat()
        manifesto.append(
            {
                "arquivo": caminho.name,
                "tamanho": stat.st_size,
                "mtime_ns": stat.st_mtime_ns,
            }
        )

    return manifesto


def indice_coluna(headers: list[str], nome: str) -> int | None:
    try:
        return headers.index(nome)
    except ValueError:
        return None


def valor_coluna(linha: tuple[Any, ...], indice: int | None) -> Any:
    if indice is None or indice >= len(linha):
        return None
    return linha[indice]


def construir_indice_lexical() -> dict[str, Any]:
    entradas: list[LexicalEntry] = []
    arquivos_indexados: dict[str, Any] = {}
    indice_tokens: dict[str, list[int]] = defaultdict(list)

    for caminho in sorted(LEXICAL_DIR.glob("*.xlsx")):
        workbook = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [str(col).strip().lower() if col is not None else "" for col in header_row]
            coluna_texto = indice_coluna(headers, "text")

            if coluna_texto is None:
                continue

            coluna_pagina = indice_coluna(headers, "pagina")
            entradas_arquivo: list[int] = []
            referencias_arquivo: list[int] = []

            for ordem, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                texto_bruto = valor_coluna(row, coluna_texto)
                if valor_ausente(texto_bruto):
                    continue

                texto = str(texto_bruto).strip()
                if not texto:
                    continue

                texto_norm = normalizar(texto)
                pagina = pagina_valida(valor_coluna(row, coluna_pagina))
                referencia_contexto = pagina if pagina is not None else ordem
                entrada = LexicalEntry(
                    arquivo=caminho.stem,
                    pagina=pagina,
                    texto=texto,
                    texto_norm=texto_norm,
                    alnum_len=contar_alnum(texto_norm),
                    ordem=ordem,
                    referencia_contexto=referencia_contexto,
                )
                indice_entrada = len(entradas)
                entradas.append(entrada)
                entradas_arquivo.append(indice_entrada)
                referencias_arquivo.append(referencia_contexto)

                for token in tokenizar(texto_norm, MAX_TOKENS_DOCUMENTO):
                    indice_tokens[token].append(indice_entrada)

            arquivos_indexados[caminho.stem] = {
                "indices": tuple(entradas_arquivo),
                "referencias": tuple(referencias_arquivo),
            }
        finally:
            workbook.close()

    return {
        "entradas": entradas,
        "arquivos": arquivos_indexados,
        "indice_tokens": {token: tuple(indices) for token, indices in indice_tokens.items()},
        "arquivos_disponiveis": sorted(arquivos_indexados),
    }


def carregar_indice_cacheado(manifesto_atual: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not INDEX_CACHE_PATH.exists():
        return None

    try:
        with INDEX_CACHE_PATH.open("rb") as arquivo_cache:
            payload = pickle.load(arquivo_cache)
    except (OSError, pickle.PickleError, EOFError, AttributeError, ValueError):
        return None

    if payload.get("version") != INDEX_CACHE_VERSION:
        return None
    if payload.get("manifesto") != manifesto_atual:
        return None

    indice = payload.get("indice")
    if not isinstance(indice, dict):
        return None

    indice["origem_indice"] = "cache_disco"
    return indice


def salvar_indice_cacheado(manifesto_atual: list[dict[str, Any]], indice: dict[str, Any]) -> None:
    payload = {
        "version": INDEX_CACHE_VERSION,
        "manifesto": manifesto_atual,
        "indice": {
            "entradas": indice["entradas"],
            "arquivos": indice["arquivos"],
            "indice_tokens": indice["indice_tokens"],
            "arquivos_disponiveis": indice["arquivos_disponiveis"],
        },
    }

    with INDEX_CACHE_PATH.open("wb") as arquivo_cache:
        pickle.dump(payload, arquivo_cache, protocol=pickle.HIGHEST_PROTOCOL)


def carregar_indice_lexical() -> dict[str, Any]:
    if not LEXICAL_DIR.exists():
        raise FileNotFoundError(f"Pasta Lexical nao encontrada em {LEXICAL_DIR}")

    manifesto_atual = coletar_manifesto_lexical()
    with PROCESS_INDEX_CACHE_LOCK:
        manifesto_cacheado = PROCESS_INDEX_CACHE.get("manifesto")
        indice_cacheado = PROCESS_INDEX_CACHE.get("indice")
        if manifesto_cacheado == manifesto_atual and isinstance(indice_cacheado, dict):
            return indice_cacheado

        indice_disco = carregar_indice_cacheado(manifesto_atual)
        if indice_disco is not None:
            PROCESS_INDEX_CACHE["manifesto"] = manifesto_atual
            PROCESS_INDEX_CACHE["indice"] = indice_disco
            return indice_disco

        indice = construir_indice_lexical()
        salvar_indice_cacheado(manifesto_atual, indice)
        indice["origem_indice"] = "reindexado"
        PROCESS_INDEX_CACHE["manifesto"] = manifesto_atual
        PROCESS_INDEX_CACHE["indice"] = indice
        return indice


def iterar_indices(indice_lexical: dict[str, Any], indices: Iterable[int] | None) -> Iterable[int]:
    if indices is None:
        return range(len(indice_lexical["entradas"]))
    return indices


def selecionar_janela_indices(
    indice_lexical: dict[str, Any],
    ultimo_resultado: dict[str, Any] | None,
    paginas_antes: int,
    paginas_depois: int,
) -> tuple[int, ...] | None:
    if ultimo_resultado is None:
        return None

    arquivo = ultimo_resultado["_arquivo"]
    referencia_base = ultimo_resultado["_referencia_contexto"]
    dados_arquivo = indice_lexical["arquivos"].get(arquivo)
    if not dados_arquivo or referencia_base is None:
        return None

    inicio = referencia_base - paginas_antes
    fim = referencia_base + paginas_depois
    referencias = dados_arquivo["referencias"]
    indice_inicio = bisect_left(referencias, inicio)
    indice_fim = bisect_right(referencias, fim)

    janela = dados_arquivo["indices"][indice_inicio:indice_fim]
    return janela or dados_arquivo["indices"]


def match_inicio(trecho_inicio: str, indice_lexical: dict[str, Any], indices: Iterable[int] | None) -> dict[str, Any]:
    if not trecho_inicio:
        return criar_resultado_vazio()

    entradas = indice_lexical["entradas"]
    for indice in iterar_indices(indice_lexical, indices):
        entrada = entradas[indice]
        if trecho_inicio in entrada.texto_norm:
            return criar_resultado(entrada, 100, "inicio")

    return criar_resultado_vazio()


def candidato_fuzzy_valido(entrada: LexicalEntry, tokens_consulta: set[str], tamanho_consulta: int) -> bool:
    if entrada.alnum_len < max(8, min(24, tamanho_consulta // 4)):
        return False

    if not tokens_consulta:
        return True

    return any(token in tokens_consulta for token in tokenizar(entrada.texto_norm, MAX_TOKENS_DOCUMENTO))


def match_fuzzy_refinado(
    trecho_fuzzy: str,
    indice_lexical: dict[str, Any],
    indices: Iterable[int] | None,
    tokens_consulta: set[str],
) -> dict[str, Any]:
    if not trecho_fuzzy:
        return criar_resultado_vazio()

    entradas = indice_lexical["entradas"]
    melhor_entrada: LexicalEntry | None = None
    melhor_score = 0.0
    tamanho_consulta = contar_alnum(trecho_fuzzy)

    for indice in iterar_indices(indice_lexical, indices):
        entrada = entradas[indice]
        if not candidato_fuzzy_valido(entrada, tokens_consulta, tamanho_consulta):
            continue

        score = fuzz.partial_ratio(
            trecho_fuzzy,
            entrada.texto_norm,
            score_cutoff=melhor_score,
        )
        if score > melhor_score:
            melhor_score = score
            melhor_entrada = entrada

    if melhor_entrada is None:
        return criar_resultado_vazio()

    return criar_resultado(melhor_entrada, melhor_score, "fuzzy_refinado")


def selecionar_candidatos_globais(trecho_fuzzy: str, indice_lexical: dict[str, Any]) -> tuple[int, ...]:
    contagem_indices: Counter[int] = Counter()
    entradas = indice_lexical["entradas"]

    for token in tokenizar(trecho_fuzzy, MAX_TOKENS_CONSULTA):
        for indice_entrada in indice_lexical["indice_tokens"].get(token, ()):
            contagem_indices[indice_entrada] += 1

    if not contagem_indices:
        return ()

    indices_ordenados = sorted(
        contagem_indices,
        key=lambda indice_entrada: (
            -contagem_indices[indice_entrada],
            abs(len(entradas[indice_entrada].texto_norm) - len(trecho_fuzzy)),
        ),
    )
    return tuple(indices_ordenados[:MAX_CANDIDATOS_GLOBAIS])


def separar_paragrafos(texto: str) -> list[str]:
    blocos = re.split(r"\n\s*\n+", texto.strip())
    paragrafos: list[str] = []

    for bloco in blocos:
        linhas = [linha.strip() for linha in bloco.splitlines() if linha.strip()]
        if not linhas:
            continue
        paragrafos.append(" ".join(linhas))

    return paragrafos


def encontrar(
    texto_original: str,
    indice_lexical: dict[str, Any],
    entradas_contexto: Iterable[int] | None,
    score_minimo_fallback: int,
) -> dict[str, Any]:
    texto_original = str(texto_original)
    if not texto_original.strip():
        return criar_resultado_vazio()

    trecho_inicio = normalizar(texto_original[:120])
    trecho_fuzzy = normalizar(texto_original[:200])
    tokens_consulta = set(tokenizar(trecho_fuzzy, MAX_TOKENS_CONSULTA))

    resultado = match_inicio(trecho_inicio, indice_lexical, entradas_contexto)
    if resultado["_arquivo"] is None:
        if entradas_contexto is None:
            candidatos_globais = selecionar_candidatos_globais(trecho_fuzzy, indice_lexical)
            resultado = match_fuzzy_refinado(trecho_fuzzy, indice_lexical, candidatos_globais, tokens_consulta)
        else:
            resultado = match_fuzzy_refinado(trecho_fuzzy, indice_lexical, entradas_contexto, tokens_consulta)

    usa_contexto = entradas_contexto is not None
    if not usa_contexto and resultado["_arquivo"] is not None:
        return resultado

    if resultado["_arquivo"] is None or (usa_contexto and resultado["score"] < score_minimo_fallback):
        candidatos_globais = selecionar_candidatos_globais(trecho_fuzzy, indice_lexical)
        resultado_global = match_inicio(trecho_inicio, indice_lexical, candidatos_globais)

        if resultado_global["_arquivo"] is None:
            resultado_global = match_fuzzy_refinado(trecho_fuzzy, indice_lexical, candidatos_globais, tokens_consulta)

        if resultado["_arquivo"] is None or resultado_global["score"] > resultado["score"]:
            return resultado_global

    return resultado


def processar_paragrafos(
    paragrafos: list[str],
    indice_lexical: dict[str, Any],
    cache_resultados: dict[str, dict[str, Any]],
    ultimo_resultado: dict[str, Any] | None,
    paginas_antes: int,
    paginas_depois: int,
    score_minimo_fallback: int,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    resultados: list[dict[str, Any]] = []
    ultimo_resultado_lote = ultimo_resultado

    for texto_original in paragrafos:
        texto_original = str(texto_original)

        if texto_original in cache_resultados:
            resultado = cache_resultados[texto_original]
        else:
            entradas_contexto = selecionar_janela_indices(
                indice_lexical,
                ultimo_resultado_lote,
                paginas_antes,
                paginas_depois,
            )
            resultado = encontrar(
                texto_original,
                indice_lexical,
                entradas_contexto,
                score_minimo_fallback,
            )
            cache_resultados[texto_original] = resultado

        resultados.append(resultado_para_saida(texto_original, resultado))

        if resultado["_arquivo"] is not None:
            ultimo_resultado_lote = resultado

    return resultados, ultimo_resultado_lote


def lookup_citations(
    text: str,
    paginas_antes: int = 2,
    paginas_depois: int = 3,
    score_minimo_fallback: int = SCORE_MINIMO_FALLBACK,
) -> dict[str, Any]:
    paragrafos = separar_paragrafos(str(text or ""))
    if not paragrafos:
        raise ValueError("Informe ao menos um paragrafo separado por linhas em branco.")

    indice_lexical = carregar_indice_lexical()
    resultados, _ultimo_resultado = processar_paragrafos(
        paragrafos,
        indice_lexical,
        {},
        None,
        max(0, int(paginas_antes)),
        max(0, int(paginas_depois)),
        int(score_minimo_fallback),
    )

    return {
        "paragraphsCount": len(paragrafos),
        "results": resultados,
        "total": len(resultados),
    }
